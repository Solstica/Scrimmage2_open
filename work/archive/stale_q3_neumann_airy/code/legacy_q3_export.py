# -*- coding: utf-8 -*-
"""把第三问 Airy 物理拟合结果整理成中文 XLSX。"""
from __future__ import annotations

import json
import math
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT = Path(__file__).resolve().parents[3]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from shared.code.data_io import load_official_bundle
from 问题3求解 import band, q2_sic_parameters, reflectance, sic_q2_reflectance


def write_book(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    book = load_workbook(path)
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(30, max(12, max(len(str(c.value)) if c.value is not None else 0 for c in column) + 2))
            sheet.column_dimensions[letter].width = width
    book.save(path)


def main() -> None:
    raw_path = PROJECT / "modules" / "40_q3" / "结果" / "问题3原始结果.json"
    if not raw_path.exists():
        raise FileNotFoundError("请先运行 问题3求解.py")
    result = json.loads(raw_path.read_text(encoding="utf-8"))
    bundle = load_official_bundle(Path(r"D:\qq文件"))
    out = PROJECT / "modules" / "40_q3" / "Origin数据"
    result_out = PROJECT / "modules" / "40_q3" / "结果"
    out.mkdir(parents=True, exist_ok=True)

    # 01: SiC 回溯，双光束和 Airy 只用于展示同一 Q2 厚度下的物理差异。
    q2 = q2_sic_parameters(PROJECT)
    sic_rows = []
    for label, s in zip(("10度单角度", "15度单角度"), bundle["sic"]):
        ss = band(s, (2500.0, 3300.0)); values = q2[label]
        double = sic_q2_reflectance(ss, values, "double"); airy = sic_q2_reflectance(ss, values, "airy")
        sic_rows.extend([[x, ss.angle_deg, 100*y, 100*d, 100*a, 100*(y-d), 100*(y-a)] for x,y,d,a in zip(ss.wavenumber_cm,ss.reflectance,double,airy)])
    write_book(out / "01_SiC回溯双光束与Airy.xlsx", {"数据": pd.DataFrame(sic_rows, columns=["波数_cm^-1","入射角_度","实测反射率_百分比","双光束预测_百分比","Airy预测_百分比","双光束残差_百分点","Airy残差_百分点"])})

    # 02: Si 光谱拟合；每个角度在独立行段内保留 X 和三条 Y 曲线。
    si_rows = []
    for s, d, a in zip((band(x,(1500.0,3500.0)) for x in bundle["si"]), result["Si"]["单角度双光束"], result["Si"]["单角度Airy"]):
        p_d=np.asarray(d["parameters"],float); p_a=np.asarray(a["parameters"],float); pred_d=reflectance(s,p_d,"double"); pred_a=reflectance(s,p_a,"airy")
        si_rows.extend([[x,s.angle_deg,100*y,100*dd,100*aa,100*(y-dd),100*(y-aa)] for x,y,dd,aa in zip(s.wavenumber_cm,s.reflectance,pred_d,pred_a)])
    write_book(out / "02_Si单角度双光束与Airy.xlsx", {"数据": pd.DataFrame(si_rows, columns=["波数_cm^-1","入射角_度","实测反射率_百分比","双光束预测_百分比","Airy预测_百分比","双光束残差_百分点","Airy残差_百分点"])})

    # 03: Si Airy 单角度与主结果，不再输出有限阶收敛表。
    rows = []
    airy = result["Si"]["单角度Airy"]
    for angle, fit in zip((10,15), airy):
        p=np.asarray(fit["parameters"],float); rows.append([angle,p[0],p[1],10**p[2],10**p[3],2*math.pi*29979245800*(10**p[3]),fit["metrics"]["RMSE_百分点"],fit["metrics"]["R2"]])
    rows.append(["平均",result["Si"]["主厚度_微米"],np.nan,np.nan,np.nan,np.nan,np.nan,np.nan])
    write_book(out / "03_Si_Airy单角度参数.xlsx", {"数据": pd.DataFrame(rows, columns=["入射角_度","Airy厚度_微米","衬底折射率","载流子浓度_cm^-3","碰撞波数_Gammae_cm^-1","碰撞角频率_Gammae_s^-1","RMSE_百分点","R2"])})

    # 04: 双角度 Airy 与双光束联合验证。
    joint = result["Si"]["联合Airy验证"]; joint_d = result["Si"]["联合双光束"]
    rows = [["联合Airy验证",joint["parameters"][0],joint["metrics"]["RMSE_百分点"],joint["metrics"]["R2"]],["联合双光束对照",joint_d["parameters"][0],joint_d["metrics"]["RMSE_百分点"],joint_d["metrics"]["R2"]],["单角度Airy平均",result["Si"]["主厚度_微米"],np.nan,np.nan]]
    write_book(out / "04_Si联合验证.xlsx", {"数据": pd.DataFrame(rows, columns=["结果类型","厚度_微米","RMSE_百分点","R2"])})

    # 05: 判据数据，适合柱状图并可加阈值线。
    criterion = pd.DataFrame(result["第三束判据"] + result["SiC"]["第三束判据"])
    write_book(out / "05_第三束光强比判据.xlsx", {"数据": criterion})

    # 06: Si 单角度双光束/Airy 对照和 RMSE 变化。
    rows=[]
    for angle,d,a in zip((10,15),result["Si"]["单角度双光束"],result["Si"]["单角度Airy"]):
        dm,am=d["metrics"],a["metrics"]; rows.append([angle,d["parameters"][0],a["parameters"][0],dm["RMSE_百分点"],am["RMSE_百分点"],100*(dm["RMSE_百分点"]-am["RMSE_百分点"])/max(dm["RMSE_百分点"],1e-30)])
    rows.append(["平均",np.mean([r[1] for r in rows]),np.mean([r[2] for r in rows]),np.mean([r[3] for r in rows]),np.mean([r[4] for r in rows]),np.mean([r[5] for r in rows])])
    write_book(out / "06_Si单角度厚度与RMSE.xlsx", {"数据": pd.DataFrame(rows, columns=["入射角_度","双光束厚度_微米","Airy厚度_微米","双光束RMSE_百分点","AiryRMSE_百分点","RMSE变化率_百分比"])})

    core = pd.DataFrame([
        ["Si","10度 Airy",airy[0]["parameters"][0],airy[0]["metrics"]["RMSE_百分点"],airy[0]["metrics"]["R2"],result["第三束判据"][0]["第三束最大光强比_百分比"],"主结果分量"],
        ["Si","15度 Airy",airy[1]["parameters"][0],airy[1]["metrics"]["RMSE_百分点"],airy[1]["metrics"]["R2"],result["第三束判据"][1]["第三束最大光强比_百分比"],"主结果分量"],
        ["Si","单角度 Airy 平均",result["Si"]["主厚度_微米"],np.nan,np.nan,np.nan,"问题三主厚度"],
        ["Si","联合 Airy 验证",joint["parameters"][0],joint["metrics"]["RMSE_百分点"],joint["metrics"]["R2"],np.nan,"仅验证"],
        ["SiC","问题二厚度回溯",result["SiC"]["问题二厚度_微米"],np.nan,np.nan,max(x["第三束最大光强比_百分比"] for x in result["SiC"]["第三束判据"]),"保留问题二结果"],
    ], columns=["材料","结果类型","厚度_微米","RMSE_百分点","R2","第三束最大光强比_百分比","用途"])
    write_book(result_out / "问题3结果汇总.xlsx", {"核心结果":core,"Si参数":pd.DataFrame(rows, columns=["入射角_度","双光束厚度_微米","Airy厚度_微米","双光束RMSE_百分点","AiryRMSE_百分点","RMSE变化率_百分比"]),"第三束判据":criterion})

    report = """# 问题三 Origin 数据使用报告（修正版）

本版正式路线是“第三束光强比判据 → 双光束或 Airy 多光束模型”。已删除有限阶 Neumann、AIC 和角度经验增益/偏置。所有文件均为 XLSX，列名已经是中文。

| 文件 | 推荐图 | Origin 操作 |
|---|---|---|
| 01_SiC回溯双光束与Airy.xlsx | 分组折线图 | 第一列设 X，按 10°/15°筛选后选择实测、双光束、Airy 三列 Y；两种模型线型不同。 |
| 02_Si单角度双光束与Airy.xlsx | 分面折线图 | 第一列 X，按入射角分组；实测用散点，双光束虚线，Airy 实线。 |
| 03_Si_Airy单角度参数.xlsx | 参数点图 | 以入射角为分类 X，厚度、RMSE 可分别作散点；不要把载流子浓度与厚度放在同一 Y 轴。 |
| 04_Si联合验证.xlsx | 分组柱状图 | 选“结果类型”和“厚度_微米”，突出单角度 Airy 平均，联合 Airy 仅作验证。 |
| 05_第三束光强比判据.xlsx | 分组柱状图 | 选材料、入射角和第三束最大光强比；在 Y=0.1 处添加红色水平线。 |
| 06_Si单角度厚度与RMSE.xlsx | 双 Y 柱状图 | 左 Y 选厚度列，右 Y 选 RMSE 列；平均行单独保留。 |

通用步骤：打开对应 XLSX，检查第一列为 X、其余数值列为 Y；在 Plot 菜单选择推荐图。光谱图横轴用波数 (cm^-1)，反射率用百分比；判据图纵轴从 0 开始，阈值固定为 0.1%。
"""
    (out / "图表设计报告.md").write_text(report, encoding="utf-8")


if __name__ == "__main__":
    main()
