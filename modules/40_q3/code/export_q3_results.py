"""Create normalized Q3 figures and editable XLSX sources from frozen JSON/CSV."""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import FancyBboxPatch
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


COLORS = {"measured": "#222222", "double": "#2F6B9A", "airy": "#D97732", "accent": "#B23A48", "green": "#4D8B62"}


def style_workbook(path: Path) -> None:
    book = load_workbook(path)
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5597")
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            width = min(34, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    book.save(path)


def write_book(path: Path, frame: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_excel(path, index=False, engine="openpyxl")
    style_workbook(path)


def model_comparison(frame: pd.DataFrame, output: Path, selected: str) -> None:
    fig, axes = plt.subplots(2, 1, figsize=(8.0, 6.7), sharex=True)
    for axis, (angle, block) in zip(axes, frame.groupby("angle_deg", sort=True)):
        axis.scatter(block["wavenumber_cm-1"], block["measured_reflectance_percent"], s=4, color=COLORS["measured"], alpha=.45, label="Measured")
        axis.plot(block["wavenumber_cm-1"], block["double_reflectance_percent"], color=COLORS["double"], lw=1.5, label="Double beam (selected)")
        axis.plot(block["wavenumber_cm-1"], block["airy_reflectance_percent"], color=COLORS["airy"], lw=1.1, ls="--", label="Airy (control)")
        axis.set_ylabel("Reflectance (%)")
        axis.set_title(f"Incidence angle {angle:.0f}°")
        axis.grid(alpha=.22)
    axes[0].legend(ncol=3, frameon=False, fontsize=8)
    axes[-1].set_xlabel(r"Wavenumber (cm$^{-1}$)")
    fig.suptitle(f"Si model comparison; criterion selects {selected} beam model")
    fig.tight_layout()
    fig.savefig(output, dpi=300, bbox_inches="tight")
    plt.close(fig)


def criterion_plot(frame: pd.DataFrame, output: Path) -> None:
    labels = [f"{m}\n{a:.0f}°" for m, a in zip(frame["material"], frame["angle_deg"])]
    values = frame["third_beam_ratio_max_percent"].to_numpy(float)
    x = np.arange(len(values))
    fig, axis = plt.subplots(figsize=(7.4, 4.5))
    colors = [COLORS["double"] if m == "Si" else COLORS["green"] for m in frame["material"]]
    axis.bar(x, values, color=colors, width=.62)
    axis.axhline(.1, color=COLORS["accent"], ls="--", lw=1.6, label="0.1% threshold")
    axis.set_yscale("log")
    axis.set_xticks(x, labels)
    axis.set_ylabel("Maximum third-beam ratio (%)")
    axis.set_title("Third-beam criterion over the prescribed spectral bands")
    axis.grid(axis="y", which="both", alpha=.22)
    axis.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(output, dpi=300, bbox_inches="tight")
    plt.close(fig)


def thickness_plot(data: dict, output: Path) -> None:
    selected = data["Si"]["selected_model"]
    source = data["Si"]["single_angle_double"] if selected == "double" else data["Si"]["single_angle_airy"]
    values = [source[0]["parameters"]["d_um"], source[1]["parameters"]["d_um"], data["Si"]["main_thickness_um"], data["Si"]["joint_selected_thickness_um"]]
    labels = ["10°", "15°", "Single-angle mean", "Joint validation"]
    fig, axis = plt.subplots(figsize=(7.4, 4.3))
    axis.plot(np.arange(4), values, "o-", color=COLORS["double"], lw=1.6, ms=7)
    axis.axhline(values[2], color=COLORS["accent"], ls="--", lw=1.0)
    axis.set_xticks(np.arange(4), labels)
    axis.set_ylabel("Si epilayer thickness (μm)")
    axis.set_title("Cross-angle and joint-fit thickness consistency")
    axis.grid(axis="y", alpha=.25)
    for i, value in enumerate(values):
        axis.text(i, value + .006, f"{value:.6f}", ha="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(output, dpi=300, bbox_inches="tight")
    plt.close(fig)


def algorithm_flow(output: Path) -> None:
    fig, axis = plt.subplots(figsize=(11.2, 3.2))
    axis.set_xlim(0, 11.2); axis.set_ylim(0, 3.2); axis.axis("off")
    boxes = [
        (0.25, "Read four spectra\nand Q2 frozen result", "#E7F0FA"),
        (2.45, "Fix 293 K dispersion;\nfit double-beam model", "#FFF1D6"),
        (4.65, "Compute max third-beam\nratio on full band", "#FCE2E6"),
        (6.85, "Select double beam if <0.1%;\notherwise select Airy", "#E8F4E8"),
        (9.05, "Average two angle estimates;\njoint fit and diagnostics", "#EAE4F5"),
    ]
    for x, label, color in boxes:
        patch = FancyBboxPatch((x, 1.05), 1.85, 1.1, boxstyle="round,pad=0.05,rounding_size=0.08", facecolor=color, edgecolor="#44546A", lw=1.2)
        axis.add_patch(patch); axis.text(x + .925, 1.60, label, ha="center", va="center", fontsize=8.5)
    for x in (2.1, 4.3, 6.5, 8.7):
        axis.annotate("", xy=(x + .33, 1.60), xytext=(x, 1.60), arrowprops=dict(arrowstyle="->", color="#44546A", lw=1.4))
    axis.text(5.6, 2.75, "Q3 conditional inversion workflow", ha="center", fontsize=12, weight="bold")
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", type=Path, required=True)
    args = parser.parse_args()
    project = args.project.resolve()
    result = json.loads((project / "output/results/q3_analysis_results.json").read_text(encoding="utf-8"))
    tables = project / "modules/40_q3/tables"
    figures = project / "modules/40_q3/figures"
    editable = figures / "editable"
    figures.mkdir(parents=True, exist_ok=True); editable.mkdir(parents=True, exist_ok=True)

    comparison = pd.read_csv(tables / "q3_si_model_comparison.csv")
    criterion = pd.read_csv(tables / "q3_third_beam_criterion.csv")
    si_results = pd.read_csv(tables / "q3_si_results.csv")
    threshold = pd.read_csv(tables / "q3_threshold_sensitivity.csv")
    sic_curve = pd.read_csv(tables / "q3_sic_criterion_curve.csv")
    for name, frame in {
        "q3_si_model_comparison.xlsx": comparison,
        "q3_third_beam_criterion.xlsx": criterion,
        "q3_si_results.xlsx": si_results,
        "q3_threshold_sensitivity.xlsx": threshold,
        "q3_sic_criterion_curve.xlsx": sic_curve,
    }.items():
        write_book(editable / name, frame)

    model_comparison(comparison, figures / "q3_si_model_comparison.png", result["Si"]["selected_model"])
    criterion_plot(criterion, figures / "q3_third_beam_criterion.png")
    thickness_plot(result, figures / "q3_thickness_consistency.png")
    algorithm_flow(figures / "q3_algorithm_flow.pdf")
    (editable / "README.md").write_text(
        "# Q3 editable figure sources\n\nEach XLSX mirrors one frozen CSV table. Import it into OriginPro; do not edit numerical values manually. Re-run `export_q3_results.py` after every model change.\n",
        encoding="utf-8", newline="\n",
    )


if __name__ == "__main__":
    main()
